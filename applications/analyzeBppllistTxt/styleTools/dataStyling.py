# External import
from openpyxl.styles import Alignment, Font, PatternFill

def data_styling(ws, row_num, li, width_rate, font_size, font_family):
    #  모든 행에 적용
    # 데이터에 개행도 모두 적용시키기 위한 작업 Alignment(wrapText=True)
    for i in li:
        #  width = 0
        max_length = 0  # 최대 길이

        for j in range(1, row_num):
            ws[i + str(j)].alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
            ws[i + str(j)].font = Font(size=font_size, bold=False, name=font_family)
            # 제일 큰 행을 기준으로 넣기
            # for k in str(ws[i + str(j)].value).split('\n'):
                # if width < len(k):
                    # width = len(k)

            cell_value = ws[i + str(j)].value
            if cell_value:
                # 문자열 길이 계산 (개행 문자로 나뉜 각 줄의 최대 길이)
                max_length = max(max_length, max(len(line) for line in str(cell_value).split('\n')))
            
        calculated_width = (max_length * width_rate)
        added_width = 6  # 추가 넓이

        # 넓이 고정
        ws.column_dimensions[i].width = calculated_width + added_width

    # print(ws["E21"].value)
    # count = 0
    # for i in ws["E21"].value:
    #     if '\n' == i:
    #         count += 1
    # print(count)


def data_row_color(ws, row_num, li, retention_name):
    pass

def data_merge_with_empty(ws, row_num, li):
    for j in range(1, len(li)):
        for i in range(1, row_num):

            # Schedule 기준 멈추기
            if ws.cell(row=i, column=j).value == "Schedule":
                return

            merge_cell_count = 0
            if ws.cell(row=i+1, column=j).value is not None \
                    and ws.cell(row=i, column=j).value is None \
                    and i + merge_cell_count < row_num:
                continue

            while ws.cell(row=i+merge_cell_count+1, column=j).value is None \
                    and i + merge_cell_count < row_num:
                merge_cell_count += 1

            # Merge 안해도 되는건 하지 않기
            if merge_cell_count != 0:
                ws.merge_cells(
                    start_row=i, start_column=j,
                    end_row=i+merge_cell_count, end_column=j
                )

                # merge시 혹시라도 낮은 높이의 셀의 경우 처리 방식
                if ws.cell(row=i, column=j).value is None :
                    continue

                count = 1
                for k in ws.cell(row=i, column=j).value:
                    if '\n' == k:
                        count += 1
                ws.row_dimensions[i].height = 17
                if ws.row_dimensions[i].height < ((count // (merge_cell_count+1)) * 16.88) + 17:
                    for l in range(i, i + merge_cell_count + 1):
                        ws.row_dimensions[l].height = round((count // (merge_cell_count+1)) * 16.88, 2) + 17

