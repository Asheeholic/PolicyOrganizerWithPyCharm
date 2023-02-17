# External import
from openpyxl.styles import Alignment

# Local import
from analyzeTools import dataRefine
from analyzeTools import putInDataInCell
from analyzeTools import putInArrayAsDict

# 정의 함수
def execute(sheet_name, ws, lines):

    # 시트 이름 (왼쪽 하단 이름)
    ws.title = sheet_name

    # 엑셀 스타일을 위한 정책 개수 리턴하기
    row_num = 2

    # 배열 선언
    rows = 0
    cols = 0
    cols_temp = 0

    for line in lines :
        line = line.strip()
        if '---------------------------------------------------------' in line :
            rows += 1
            cols_temp = 0

        cols_temp += 1
        if cols <= cols_temp :
            cols = cols_temp

    analyze_txt_lines = [["HANJAEHO" for a in range(cols)] for b in range(rows + 1)]

    line_i = 0
    line_j = 0


    # 2차원 배열 집어넣기
    for line in lines :
        line = line.strip()
        analyze_txt_lines[line_i][line_j] = line

        line_j += 1
        if line == '------------------------------------------------------------' :
            line_i += 1
            line_j = 0


    ### 엑셀에 집어 넣기

    # 1. 열 별 제목 만들기
    column_titles = [
        'Policy Name',
        'Policy Type',
        'Active',
        'Residence',  # storage unit
        'HW/OS/Client',
        'Include',  # backup selection
        'Schedule',  # Schedule name
        'Type',  # schedule type
        'Other Residence', # Copy 2 and more
        'Retention Level',  # retention (레벨은 안찍히고 뒤에 날짜만)
        'Frequency',  # Frequency / calendar
        'Calendar sched',
        'Daily Windows'
    ]
    title_name_i = 1
    for title_name in column_titles :
        ws.cell(column=title_name_i, row=1, value=title_name)
        title_name_i += 1

    # 2. 정책 별 정보 넣기

    # 정책 번호

    # print(analyze_txt_lines[3])
    ## 들어가면 안되는 문자들
    not_in_the_list = [
        'HANJAEHO',
        'No specific exclude dates entered',
        'No exclude days of week entered',
        'Excluded Dates----------',
        '',
        '------------------------------------------------------------'
    ]

    for policy_num in range(1, len(analyze_txt_lines)):

        datas = putInArrayAsDict.put_in_array_as_dict(analyze_txt_lines[policy_num], not_in_the_list)
        # refine
        datas = dataRefine.data_refine('HW/OS/Client', datas)
        datas = dataRefine.data_refine('Include', datas)
        datas = dataRefine.data_refine('Daily Windows', datas)
        datas = dataRefine.residence_refine(datas)
        datas = dataRefine.retention_refine(datas)
        datas = dataRefine.calendar_refine(datas)
        datas = dataRefine.daily_windows_refine_for_7days(datas)
        datas = dataRefine.none_refine(datas)
        row_num = putInDataInCell.put_in_data_in_cell(ws, column_titles, datas, row_num)

    return row_num

    # #### 한 셀 줄바꿈 방법
    # ws['A1'] = "Line 1\nLine 2"
    # ws['A1'].alignment = Alignment(wrap_text=True) # 하나의 셀에 여러 행을 넣을 수 있게 만듬.
