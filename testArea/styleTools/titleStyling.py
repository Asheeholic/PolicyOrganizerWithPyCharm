# External import
from openpyxl.styles import Alignment, Font, PatternFill

def title_styling(ws, li):
    # 타이틀 만들기
    for i in range(0, len(li)) :
        li[i] = li[i] + "1"

    # 타이틀 꾸미기
    for item in li :
        ws[item].alignment = Alignment(vertical='center', wrap_text=True)
        ws[item].font = Font(size=11, bold=True)
        ws[item].fill = PatternFill(start_color = 'F4B084', end_color = 'F4B084', patternType='solid')

        if str(ws[item].value) == 'Residence':
            ws[item].value = 'Storage Unit'

        elif str(ws[item].value) == 'Include':
            ws[item].value = 'Backup Selection'
