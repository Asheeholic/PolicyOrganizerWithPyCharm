# External import
from openpyxl.styles import Alignment, Font, PatternFill

def data_styling(ws, row_num, li, width_rate):
    #  모든 행에 적용
    for i in li:
        width = 0
        for j in range(1, row_num):
            ws[i + str(j)].alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
            # 제일 큰 행을 기준으로 넣기
            for k in str(ws[i + str(j)].value).split('\n'):
                if width < len(k):
                    width = len(k)

        # 넓이 고정
        ws.column_dimensions[i].width = (width * width_rate) + 4

    # print(ws["E21"].value)
    # count = 0
    # for i in ws["E21"].value:
    #     if '\n' == i:
    #         count += 1
    # print(count)


def data_row_color(ws, row_num, li, retention_name):
    pass

def data_merge_with_empty(ws, row_num, li):
    for j in range(1, len(li)):
        for i in range(1, row_num):

            # Schedule 기준 멈추기
            if ws.cell(row=i, column=j).value == "Schedule":
                return

            merge_cell_count = 0
            if ws.cell(row=i+1, column=j).value is not None \
                    and ws.cell(row=i, column=j).value is None \
                    and i + merge_cell_count < row_num:
                continue

            while ws.cell(row=i+merge_cell_count+1, column=j).value is None \
                    and i + merge_cell_count < row_num:
                merge_cell_count += 1

            # Merge 안해도 되는건 하지 않기
            if merge_cell_count != 0:
                ws.merge_cells(
                    start_row=i, start_column=j,
                    end_row=i+merge_cell_count, end_column=j
                )

                # merge시 혹시라도 낮은 높이의 셀의 경우 처리 방식
                if ws.cell(row=i, column=j).value is None :
                    continue

                count = 1
                for k in ws.cell(row=i, column=j).value:
                    if '\n' == k:
                        count += 1
                ws.row_dimensions[i].height = 17
                if ws.row_dimensions[i].height < ((count // (merge_cell_count+1)) * 16.88) + 17:
                    for l in range(i, i + merge_cell_count + 1):
                        ws.row_dimensions[l].height = round((count // (merge_cell_count+1)) * 16.88, 2) + 17

